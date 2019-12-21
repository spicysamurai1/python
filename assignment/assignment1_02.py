import os
import openpyxl    

def search1(val, col):
    print()
    activate = 0
    print("Contact Details : ")
    for i in range(1,sheet.max_row+1):
        if(sheet.cell(row=i, column=col).value == val):
            print("\nName : ",sheet.cell(row=i,column=1).value,"\nPhone Number : ",sheet.cell(row=i,column=2).value,"\nEmail Address : ",sheet.cell(row=i,column=3).value,"\nHome Address : ",sheet.cell(row=i,column=4).value)
            activate = 1
    if(activate == 0):
        print("No Contact Details Found.")
def search():
    ch1 = 0
    while(ch1 != 4):
        print("\n1. Search By Name\n2. Search By Phone Number\n3. Search By Email Address.\n4. Exit")
        ch1 = int(input("Enter choice : "))
        if(ch1 == 1):
            name = input("\nEnter Name : ")
            search1(name, 1)
        elif(ch1 == 2):
            phone = int(input("\nEnter Phone Number : "))
            search1(phone, 2)
        elif(ch1 == 3):
            email = input("\nEnter Email Address : ")
            search1(email, 3)
        elif(ch1 == 4):
            pass
        else:
            print("Enter Correct Choice.")

curr_dir = os.path.dirname(__file__)
filename = os.path.join(curr_dir, "contact_book.xlsx")

if os.path.isfile(filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb['Contact_Details']
else:
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Contact_Details"
    sheet['A1'] = "Name"
    sheet['B1'] = "Phone Number"
    sheet['C1'] = "Email Address"
    sheet['D1'] = "Home Address"
    wb.save(filename)

ch = 0
while(ch != 3):
    print("\n1. Enter Contact Details\n2. Search Contact Details\n3. Exit")
    ch = int(input("Enter your choice : "))
    if(ch == 1):
        print()
        mr = sheet.max_row
        name = input("Enter Name : ")
        sheet.cell(row=mr+1, column=1, value=name)
        phone = int(input("Enter Phone Number : "))
        sheet.cell(row=mr+1, column=2, value=phone)
        email = input("Enter Email Address : ")
        sheet.cell(row=mr+1, column=3, value=email)
        home = input("Enter Home Address : ")
        sheet.cell(row=mr+1, column=4, value=home)
        wb.save(filename)
        
    elif(ch == 2):
        search()
    elif(ch == 3):
        pass
    else:
        print("Enter Correct Choice.")

